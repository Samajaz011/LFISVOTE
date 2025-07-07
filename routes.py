from flask import render_template, request, redirect, url_for, flash, session, jsonify, make_response
from app import app, db
from models import Admin, Student, Candidate, Vote, VotingSession
from werkzeug.security import generate_password_hash
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from datetime import datetime

@app.route('/')
def index():
    """Landing page"""
    active_session = VotingSession.query.filter_by(is_active=True).first()
    voting_status = get_voting_status()
    return render_template('index.html', active_session=active_session, voting_status=voting_status)

@app.route('/voting')
def voting():
    """Main voting interface"""
    active_session = VotingSession.query.filter_by(is_active=True).first()
    if not active_session:
        flash('No active voting session found.', 'warning')
        return redirect(url_for('index'))
    
    # Get voting status and time information
    voting_status = get_voting_status()
    
    head_boy_candidates = Candidate.query.filter_by(position='head_boy', is_active=True).all()
    head_girl_candidates = Candidate.query.filter_by(position='head_girl', is_active=True).all()
    
    return render_template('voting.html', 
                         head_boy_candidates=head_boy_candidates,
                         head_girl_candidates=head_girl_candidates,
                         active_session=active_session,
                         voting_status=voting_status)

def is_voting_active():
    """Check if voting is currently active based on admin control only"""
    # Check admin override settings
    voting_session = VotingSession.query.first()
    if voting_session and voting_session.admin_override:
        if voting_session.override_status == 'force_closed':
            return False
    
    # Voting is always active unless manually closed by admin
    return True

def get_voting_status():
    """Get current voting status information"""
    voting_session = VotingSession.query.first()
    return {
        'status': 'active' if is_voting_active() else 'closed',
        'message': 'Voting is open' if is_voting_active() else 'Voting is closed by admin',
        'admin_controlled': voting_session and voting_session.admin_override,
        'override_status': voting_session.override_status if voting_session and voting_session.admin_override else 'open'
    }

@app.route('/submit_vote', methods=['POST'])
def submit_vote():
    """Process vote submission"""
    # Check if voting is active
    if not is_voting_active():
        voting_status = get_voting_status()
        flash(voting_status['message'], 'error')
        return redirect(url_for('voting'))
    
    admission_number = request.form.get('student_id', '').strip()
    head_boy_candidate_id = request.form.get('head_boy_candidate')
    head_girl_candidate_id = request.form.get('head_girl_candidate')
    
    # Validate admission number
    if not admission_number:
        flash('Please enter your admission number.', 'error')
        return redirect(url_for('voting'))
    
    # Check if this admission number has already voted
    existing_student = Student.query.filter_by(student_id=admission_number).first()
    if existing_student and existing_student.has_voted:
        flash('This admission number has already been used to vote. Only one vote per admission number is allowed.', 'error')
        return redirect(url_for('voting'))
    
    # Validate candidates
    if not head_boy_candidate_id or not head_girl_candidate_id:
        flash('Please select both Head Boy and Head Girl candidates.', 'error')
        return redirect(url_for('voting'))
    
    head_boy_candidate = Candidate.query.get(head_boy_candidate_id)
    head_girl_candidate = Candidate.query.get(head_girl_candidate_id)
    
    if not head_boy_candidate or not head_girl_candidate:
        flash('Invalid candidate selection.', 'error')
        return redirect(url_for('voting'))
    
    try:
        # Create or get student record
        if not existing_student:
            student = Student(
                student_id=admission_number,
                name=admission_number,  # Store actual admission number
                has_voted=True
            )
            db.session.add(student)
            db.session.flush()  # Get the student ID
        else:
            student = existing_student
            student.has_voted = True
        
        # Create votes
        head_boy_vote = Vote(
            student_id=student.id,
            candidate_id=head_boy_candidate.id,
            position='head_boy'
        )
        
        head_girl_vote = Vote(
            student_id=student.id,
            candidate_id=head_girl_candidate.id,
            position='head_girl'
        )
        
        # Save to database
        db.session.add(head_boy_vote)
        db.session.add(head_girl_vote)
        db.session.commit()
        
        # Store vote info in session for success page
        session['vote_success'] = {
            'admission_number': admission_number,
            'head_boy': head_boy_candidate.name,
            'head_girl': head_girl_candidate.name,
            'vote_date': datetime.now()
        }
        
        return redirect(url_for('vote_success'))
        
    except Exception as e:
        db.session.rollback()
        flash('An error occurred while processing your vote. Please try again.', 'error')
        app.logger.error(f"Vote submission error: {str(e)}")
        return redirect(url_for('voting'))

@app.route('/vote_success')
def vote_success():
    """Vote success confirmation"""
    vote_info = session.get('vote_success')
    if not vote_info:
        return redirect(url_for('index'))
    
    # Clear the session data
    session.pop('vote_success', None)
    
    return render_template('vote_success.html', vote_info=vote_info)

@app.route('/admin')
def admin():
    """Admin dashboard"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    # Get statistics
    total_students = Student.query.count()
    total_votes = Vote.query.count()
    voted_students = Student.query.filter_by(has_voted=True).count()
    
    head_boy_candidates = Candidate.query.filter_by(position='head_boy', is_active=True).all()
    head_girl_candidates = Candidate.query.filter_by(position='head_girl', is_active=True).all()
    
    # Get vote counts
    head_boy_results = []
    for candidate in head_boy_candidates:
        head_boy_results.append({
            'candidate': candidate,
            'votes': Vote.query.filter_by(candidate_id=candidate.id).count()
        })
    
    head_girl_results = []
    for candidate in head_girl_candidates:
        head_girl_results.append({
            'candidate': candidate,
            'votes': Vote.query.filter_by(candidate_id=candidate.id).count()
        })
    
    # Get voting session info for admin control
    voting_session = VotingSession.query.first()
    voting_status = get_voting_status()
    current_voting_active = is_voting_active()
    
    return render_template('admin.html',
                         total_students=total_students,
                         total_votes=total_votes,
                         voted_students=voted_students,
                         head_boy_results=head_boy_results,
                         head_girl_results=head_girl_results,
                         voting_session=voting_session,
                         voting_status=voting_status,
                         current_voting_active=current_voting_active)

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Admin login"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        admin = Admin.query.filter_by(username=username).first()
        if admin and admin.check_password(password):
            session['admin_logged_in'] = True
            session['admin_username'] = username
            return redirect(url_for('admin'))
        else:
            flash('Invalid username or password.', 'error')
    
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    """Admin logout"""
    session.pop('admin_logged_in', None)
    session.pop('admin_username', None)
    return redirect(url_for('index'))

@app.route('/admin/candidates')
def admin_candidates():
    """Manage candidates"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    candidates = Candidate.query.all()
    return render_template('admin_candidates.html', candidates=candidates)

@app.route('/admin/students')
def admin_students():
    """Manage students"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    students = Student.query.all()
    return render_template('admin_students.html', students=students)

@app.route('/admin/results')
def admin_results():
    """View detailed results"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    # Get all votes with details
    votes = db.session.query(Vote, Student, Candidate).join(Student).join(Candidate).all()
    
    return render_template('admin_results.html', votes=votes)

@app.route('/admin/export_results')
def export_results():
    """Export results to Excel"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    
    # Summary headers
    ws_summary['A1'] = "Position"
    ws_summary['B1'] = "Candidate"
    ws_summary['C1'] = "Class"
    ws_summary['D1'] = "Votes"
    
    for cell in ws_summary['A1:D1'][0]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Head Boy results
    row = 2
    head_boy_candidates = Candidate.query.filter_by(position='head_boy', is_active=True).all()
    for candidate in head_boy_candidates:
        vote_count = Vote.query.filter_by(candidate_id=candidate.id).count()
        ws_summary[f'A{row}'] = "Head Boy"
        ws_summary[f'B{row}'] = candidate.name
        ws_summary[f'C{row}'] = candidate.class_name
        ws_summary[f'D{row}'] = vote_count
        row += 1
    
    # Head Girl results
    head_girl_candidates = Candidate.query.filter_by(position='head_girl', is_active=True).all()
    for candidate in head_girl_candidates:
        vote_count = Vote.query.filter_by(candidate_id=candidate.id).count()
        ws_summary[f'A{row}'] = "Head Girl"
        ws_summary[f'B{row}'] = candidate.name
        ws_summary[f'C{row}'] = candidate.class_name
        ws_summary[f'D{row}'] = vote_count
        row += 1
    
    # Detailed votes sheet
    ws_detailed = wb.create_sheet("Detailed Votes")
    
    # Detailed headers
    ws_detailed['A1'] = "Vote ID"
    ws_detailed['B1'] = "Admission Number"
    ws_detailed['C1'] = "Position"
    ws_detailed['D1'] = "Candidate"
    ws_detailed['E1'] = "Candidate Class"
    ws_detailed['F1'] = "Vote Time"
    
    for cell in ws_detailed['A1:F1'][0]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Detailed vote data
    votes = db.session.query(Vote, Student, Candidate).join(Student).join(Candidate).all()
    row = 2
    for vote, student, candidate in votes:
        ws_detailed[f'A{row}'] = vote.id
        ws_detailed[f'B{row}'] = student.student_id
        ws_detailed[f'C{row}'] = vote.position.replace('_', ' ').title()
        ws_detailed[f'D{row}'] = candidate.name
        ws_detailed[f'E{row}'] = candidate.class_name
        ws_detailed[f'F{row}'] = vote.created_at.strftime('%Y-%m-%d %H:%M:%S')
        row += 1
    
    # Auto-adjust column widths
    for ws in [ws_summary, ws_detailed]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    response = make_response(excel_file.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=voting_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

@app.route('/admin/add_candidate', methods=['POST'])
def add_candidate():
    """Add new candidate"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    name = request.form.get('name')
    class_name = request.form.get('class_name')
    position = request.form.get('position')
    description = request.form.get('description')
    achievements = request.form.get('achievements')
    
    # Handle photo upload
    photo_url = None
    if 'candidate_photo' in request.files:
        photo_file = request.files['candidate_photo']
        if photo_file and photo_file.filename:
            import os
            import uuid
            from werkzeug.utils import secure_filename
            
            # Create unique filename
            file_extension = photo_file.filename.rsplit('.', 1)[1].lower() if '.' in photo_file.filename else 'jpg'
            unique_filename = f"{uuid.uuid4()}_{secure_filename(photo_file.filename)}"
            
            # Ensure uploads directory exists
            upload_dir = os.path.join(app.static_folder, 'uploads', 'candidates')
            os.makedirs(upload_dir, exist_ok=True)
            
            # Save file
            file_path = os.path.join(upload_dir, unique_filename)
            photo_file.save(file_path)
            
            # Store relative URL for database
            photo_url = f"/static/uploads/candidates/{unique_filename}"
    
    candidate = Candidate(
        name=name,
        class_name=class_name,
        position=position,
        photo_url=photo_url,
        description=description,
        achievements=achievements
    )
    
    db.session.add(candidate)
    db.session.commit()
    
    flash('Candidate added successfully!', 'success')
    return redirect(url_for('admin_candidates'))

@app.route('/admin/edit_candidate', methods=['POST'])
def edit_candidate():
    """Edit existing candidate"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    candidate_id = request.form.get('candidate_id')
    candidate = Candidate.query.get_or_404(candidate_id)
    
    candidate.name = request.form.get('name')
    candidate.class_name = request.form.get('class_name')
    candidate.position = request.form.get('position')
    candidate.description = request.form.get('description')
    candidate.achievements = request.form.get('achievements')
    candidate.is_active = bool(int(request.form.get('is_active', 1)))
    
    # Handle photo upload
    if 'candidate_photo' in request.files:
        photo_file = request.files['candidate_photo']
        if photo_file and photo_file.filename:
            import os
            import uuid
            from werkzeug.utils import secure_filename
            
            # Delete old photo if it exists and is a file (not base64)
            if candidate.photo_url and candidate.photo_url.startswith('/static/uploads/'):
                old_file_path = os.path.join(app.static_folder, candidate.photo_url[8:])  # Remove '/static/' prefix
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)
            
            # Create unique filename
            file_extension = photo_file.filename.rsplit('.', 1)[1].lower() if '.' in photo_file.filename else 'jpg'
            unique_filename = f"{uuid.uuid4()}_{secure_filename(photo_file.filename)}"
            
            # Ensure uploads directory exists
            upload_dir = os.path.join(app.static_folder, 'uploads', 'candidates')
            os.makedirs(upload_dir, exist_ok=True)
            
            # Save file
            file_path = os.path.join(upload_dir, unique_filename)
            photo_file.save(file_path)
            
            # Store relative URL for database
            candidate.photo_url = f"/static/uploads/candidates/{unique_filename}"
    
    db.session.commit()
    
    flash('Candidate updated successfully!', 'success')
    return redirect(url_for('admin_candidates'))

@app.route('/admin/get_candidate/<int:candidate_id>')
def get_candidate(candidate_id):
    """Get candidate data for editing"""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Not authorized'}), 401
    
    candidate = Candidate.query.get_or_404(candidate_id)
    
    return jsonify({
        'id': candidate.id,
        'name': candidate.name,
        'class_name': candidate.class_name,
        'position': candidate.position,
        'photo_url': candidate.photo_url or '',
        'description': candidate.description or '',
        'achievements': candidate.achievements or '',
        'is_active': candidate.is_active
    })

@app.route('/admin/delete_candidate/<int:candidate_id>', methods=['POST'])
def delete_candidate(candidate_id):
    """Delete candidate"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    try:
        candidate = Candidate.query.get_or_404(candidate_id)
        
        # Delete photo file if it exists and is a file (not base64)
        if candidate.photo_url and candidate.photo_url.startswith('/static/uploads/'):
            import os
            file_path = os.path.join(app.static_folder, candidate.photo_url[8:])  # Remove '/static/' prefix
            if os.path.exists(file_path):
                os.remove(file_path)
        
        # Delete associated votes first
        Vote.query.filter_by(candidate_id=candidate_id).delete()
        
        # Delete the candidate
        db.session.delete(candidate)
        db.session.commit()
        
        flash('Candidate deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error deleting candidate. Please try again.', 'error')
        app.logger.error(f"Delete candidate error: {str(e)}")
    
    return redirect(url_for('admin_candidates'))

@app.route('/admin/add_student', methods=['POST'])
def add_student():
    """Add new student"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    student_id = request.form.get('student_id')
    name = request.form.get('name')
    class_name = request.form.get('class_name')
    
    student = Student(
        student_id=student_id,
        name=name,
        class_name=class_name
    )
    
    db.session.add(student)
    db.session.commit()
    
    flash('Student added successfully!', 'success')
    return redirect(url_for('admin_students'))

@app.route('/admin/reset_votes', methods=['POST'])
def reset_votes():
    """Reset all votes for next election"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    try:
        # Delete all votes
        Vote.query.delete()
        
        # Reset all students' voting status
        Student.query.update({Student.has_voted: False})
        
        db.session.commit()
        flash('All votes have been reset successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error resetting votes. Please try again.', 'error')
        app.logger.error(f"Reset votes error: {str(e)}")
    
    return redirect(url_for('admin'))

@app.route('/admin/voting_control/<action>', methods=['POST'])
def voting_control(action):
    """Admin control for starting/stopping voting"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    try:
        voting_session = VotingSession.query.first()
        if not voting_session:
            flash('No voting session found. Please contact system administrator.', 'error')
            return redirect(url_for('admin'))
        
        if action == 'force_open':
            voting_session.admin_override = True
            voting_session.override_status = 'force_open'
            flash('Voting has been manually opened! Students can now vote.', 'success')
            
        elif action == 'force_close':
            voting_session.admin_override = True
            voting_session.override_status = 'force_closed'
            flash('Voting has been manually closed! No more votes will be accepted.', 'warning')
            
        elif action == 'reset_schedule':
            voting_session.admin_override = False
            voting_session.override_status = 'scheduled'
            flash('Voting control returned to default open status.', 'info')
            
        else:
            flash('Invalid action requested.', 'error')
            return redirect(url_for('admin'))
        
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        flash('Error updating voting status. Please try again.', 'error')
        app.logger.error(f"Voting control error: {str(e)}")
    
    return redirect(url_for('admin'))

@app.route('/admin/system/credentials', methods=['GET', 'POST'])
def admin_credentials():
    """Hidden admin credentials management portal"""
    if request.method == 'POST':
        current_username = request.form.get('current_username')
        current_password = request.form.get('current_password')
        new_username = request.form.get('new_username')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        # Validate current credentials
        admin = Admin.query.filter_by(username=current_username).first()
        if not admin or not admin.check_password(current_password):
            flash('Invalid current credentials', 'error')
            return render_template('admin_credentials.html')
        
        # Validate new password
        if new_password != confirm_password:
            flash('New passwords do not match', 'error')
            return render_template('admin_credentials.html')
        
        if len(new_password) < 6:
            flash('New password must be at least 6 characters long', 'error')
            return render_template('admin_credentials.html')
        
        try:
            # Update credentials
            admin.username = new_username
            admin.set_password(new_password)
            db.session.commit()
            
            flash('Admin credentials updated successfully!', 'success')
            app.logger.info(f"Admin credentials updated - new username: {new_username}")
            return render_template('admin_credentials.html', success=True)
            
        except Exception as e:
            db.session.rollback()
            flash('Error updating credentials. Please try again.', 'error')
            app.logger.error(f"Credentials update error: {str(e)}")
    
    return render_template('admin_credentials.html')

# Initialize default admin if none exists
def create_default_admin():
    """Create default admin user if none exists"""
    if not Admin.query.first():
        admin = Admin(username='admin')
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()
        app.logger.info("Default admin created - username: admin, password: admin123")

def initialize_candidates():
    """Initialize default candidates if none exist"""
    if not Candidate.query.first():
        # Head Boy candidates
        head_boy_candidates = [
            {
                'name': 'Aditya Yadav',
                'class_name': 'XI Gamma',
                'position': 'head_boy',
                'description': 'Student leader with exceptional academic performance and strong leadership qualities.',
                'achievements': 'School Captain 2024, Academic Excellence Award, Debate Team Leader'
            },
            {
                'name': 'Arjun Kumar',
                'class_name': 'XI Beta',
                'position': 'head_boy',
                'description': 'Active in sports and community service with proven organizational skills.',
                'achievements': 'Cricket Team Captain, Community Service Award, Student Council Member'
            },
            {
                'name': 'Vinyak Singh',
                'class_name': 'XI Gamma',
                'position': 'head_boy',
                'description': 'Dedicated student with strong communication skills and innovative ideas.',
                'achievements': 'Science Fair Winner, Drama Club President, Inter-school Debate Champion'
            },
            {
                'name': 'Vikas Chauhan',
                'class_name': 'XI Beta',
                'position': 'head_boy',
                'description': 'Committed to student welfare and academic excellence with leadership experience.',
                'achievements': 'Mathematics Olympiad Gold Medal, Student Mentor, Environmental Club Leader'
            }
        ]
        
        # Head Girl candidates
        head_girl_candidates = [
            {
                'name': 'Umra Shahid',
                'class_name': 'XI Beta',
                'position': 'head_girl',
                'description': 'Outstanding academic achiever with strong interpersonal skills and leadership vision.',
                'achievements': 'Head Girl 2024, Academic Excellence Award, Cultural Committee Head'
            },
            {
                'name': 'Aakriti Yadav',
                'class_name': 'XI Alpha',
                'position': 'head_girl',
                'description': 'Creative and innovative student leader with excellent communication abilities.',
                'achievements': 'Art Competition Winner, Literary Society President, Student Voice Representative'
            },
            {
                'name': 'Priyanshi Chauhan',
                'class_name': 'XI Beta',
                'position': 'head_girl',
                'description': 'Compassionate leader focused on student welfare and academic improvement.',
                'achievements': 'Volunteer of the Year, Music Club Leader, Academic Support Coordinator'
            },
            {
                'name': 'Neha Yadav',
                'class_name': 'XI Beta',
                'position': 'head_girl',
                'description': 'Dynamic student with proven track record in organizing events and leading initiatives.',
                'achievements': 'Event Management Award, Dance Team Captain, Peer Counselor'
            }
        ]
        
        # Add all candidates to database
        all_candidates = head_boy_candidates + head_girl_candidates
        for candidate_data in all_candidates:
            candidate = Candidate(**candidate_data)
            db.session.add(candidate)
        
        # Create a default voting session
        voting_session = VotingSession(
            title="Little Flower International School Elections 2025",
            description="Vote for your preferred Head Boy and Head Girl candidates. Make your voice heard in shaping the future leadership of our school.",
            is_active=True
        )
        db.session.add(voting_session)
        
        db.session.commit()
        app.logger.info("Default candidates and voting session created")

def initialize_sample_students():
    """Initialize sample students for testing"""
    if not Student.query.first():
        sample_students = [
            {'student_id': 'S001', 'name': 'John Smith', 'class_name': 'XI Alpha'},
            {'student_id': 'S002', 'name': 'Sarah Johnson', 'class_name': 'XI Beta'},
            {'student_id': 'S003', 'name': 'Mike Wilson', 'class_name': 'XI Gamma'},
            {'student_id': 'S004', 'name': 'Emily Davis', 'class_name': 'XI Alpha'},
            {'student_id': 'S005', 'name': 'David Brown', 'class_name': 'XI Beta'},
            {'student_id': 'TEST123', 'name': 'Test Student', 'class_name': 'XI Test'},
        ]
        
        for student_data in sample_students:
            student = Student(**student_data)
            db.session.add(student)
        
        db.session.commit()
        app.logger.info("Sample students created for testing")

def initialize_voting_session():
    """Initialize voting session for election"""
    if not VotingSession.query.filter_by(is_active=True).first():
        from datetime import datetime
        voting_session = VotingSession(
            title="Head Boy and Head Girl Elections 2025",
            description="Annual student elections for school leadership positions",
            is_active=True,
            start_time=datetime(2025, 7, 7, 9, 0, 0),  # July 7, 2025, 9:00 AM
            end_time=datetime(2025, 7, 7, 14, 0, 0),   # July 7, 2025, 2:00 PM
            admin_override=False,
            override_status='scheduled'
        )
        db.session.add(voting_session)
        db.session.commit()

# Template filters
@app.template_filter('datetime')
def datetime_filter(datetime_obj):
    """Format datetime for templates"""
    if datetime_obj:
        return datetime_obj.strftime('%Y-%m-%d %H:%M:%S')
    return ''

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
